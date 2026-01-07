"""
Participation Report Generator

This module generates a participation report from an Excel file.
It creates pivot tables and bar charts showing participation by department and test status.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def read_csv_with_encoding(file_path):
    """
    Read CSV file with automatic encoding detection and robust parsing.
    Tries multiple encodings and delimiters to handle various CSV formats.
    
    Args:
        file_path (str): Path to the CSV file
    
    Returns:
        pandas.DataFrame: DataFrame with CSV data
    
    Raises:
        ValueError: If file cannot be read with any method
    """
    encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252', 'windows-1252']
    delimiters = [',', ';', '\t', '|']
    
    # Try different combinations of encoding and delimiter
    for encoding in encodings:
        for delimiter in delimiters:
            try:
                df = pd.read_csv(
                    file_path,
                    encoding=encoding,
                    delimiter=delimiter,
                    quotechar='"',
                    on_bad_lines='skip',  # Skip malformed lines
                    engine='python',  # Use Python engine for better error handling
                    skipinitialspace=True
                )
                # If we got here and have data, return it
                if len(df) > 0:
                    return df
            except (UnicodeDecodeError, UnicodeError):
                continue
            except Exception as e:
                # If it's not an encoding error, try next delimiter
                if 'codec' in str(e).lower() or 'decode' in str(e).lower():
                    continue
                # For other errors, try next delimiter
                continue
    
    # If all combinations failed, try with most lenient settings (auto-detect separator)
    for encoding in encodings:
        try:
            df = pd.read_csv(
                file_path,
                encoding=encoding,
                on_bad_lines='skip',
                engine='python',
                sep=None,  # Auto-detect separator
                skipinitialspace=True
            )
            if len(df) > 0:
                return df
        except Exception:
            continue
    
    # Last resort: try with errors='replace'
    try:
        return pd.read_csv(
            file_path,
            encoding='utf-8',
            errors='replace',
            on_bad_lines='skip',
            engine='python',
            sep=None,
            skipinitialspace=True
        )
    except Exception as e:
        raise ValueError(f"Unable to read CSV file. Error: {str(e)}")


def generate_participation_report(input_file, output_file):
    """
    Generate a participation report with pivot tables and charts.
    
    Args:
        input_file (str): Path to the input file (CSV, XLSX, or XLS)
        output_file (str): Path to save the output Excel file
    
    Returns:
        str: Path to the generated output file
    """
    # Detect file extension and read accordingly
    file_ext = input_file.lower().split('.')[-1]
    
    if file_ext == "csv":
        df = read_csv_with_encoding(input_file)
    elif file_ext in ["xlsx", "xls"]:
        try:
            df = pd.read_excel(input_file)
        except Exception as e:
            error_msg = str(e).lower()
            # Check if it's a corrupted/invalid Excel file - try CSV as fallback
            if 'cannot be used in worksheets' in error_msg or 'badzipfile' in error_msg or 'corrupt' in error_msg:
                # Try reading as CSV instead (file might be misnamed)
                try:
                    df = read_csv_with_encoding(input_file)
                    # Successfully read as CSV, continue processing
                except:
                    raise ValueError(f"Invalid or corrupted file. The file '{input_file}' cannot be read as Excel or CSV. Please ensure the file is valid.")
            else:
                raise ValueError(f"Error reading Excel file: {str(e)}. Please ensure the file is valid.")
    else:
        raise ValueError("Unsupported file format. Please upload CSV or Excel files.")
    
    # Create a new workbook
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Data"
    
    # Write data to sheet
    _write_data_to_sheet(ws_data, df)
    
    # Apply styling to Data sheet
    _style_data_sheet(ws_data, df)
    
    # Create pivot table
    pivot_df = _create_participation_pivot(df)
    
    # Create Participation Summary sheet
    ws_summary = wb.create_sheet("Participation Summary")
    _write_pivot_to_sheet(ws_summary, pivot_df, "Participation Summary")
    
    # Add bar chart
    _add_participation_chart(ws_summary, pivot_df)
    
    # Save workbook
    wb.save(output_file)
    return output_file


def _write_data_to_sheet(ws, df):
    """
    Write DataFrame data to worksheet.
    
    Args:
        ws: openpyxl worksheet object
        df: pandas DataFrame
    """
    # Write headers
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Write data
    for row_idx, row_data in enumerate(df.values, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def _style_data_sheet(ws, df):
    """
    Apply styling to the data sheet.
    
    Args:
        ws: openpyxl worksheet object
        df: pandas DataFrame
    """
    # Header styling
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Apply data styling and borders
    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            # Alternate row coloring
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Auto-adjust column widths
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_length = max(
            len(str(col_name)),
            df[col_name].astype(str).map(len).max() if len(df) > 0 else 0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)


def _create_participation_pivot(df):
    """
    Create a pivot table for participation report.
    
    Args:
        df: pandas DataFrame
    
    Returns:
        pandas DataFrame: Pivot table
    """
    # Find Department and Test Status columns (case-insensitive)
    dept_col = None
    status_col = None
    
    for col in df.columns:
        col_lower = str(col).lower()
        if 'department' in col_lower:
            dept_col = col
        if 'test status' in col_lower or 'status' in col_lower:
            status_col = col
    
    if dept_col is None or status_col is None:
        raise ValueError("Required columns 'Department' and 'Test Status' not found in the Excel file")
    
    # Find Name column for counting (or use first column)
    name_col = None
    for col in df.columns:
        if 'name' in str(col).lower():
            name_col = col
            break
    
    if name_col is None:
        # Use first column as fallback
        name_col = df.columns[0]
    
    # Create pivot table with Grand Total row and column
    pivot_df = pd.pivot_table(
        df,
        index=dept_col,
        columns=status_col,
        values=name_col,
        aggfunc='count',
        fill_value=0,
        margins=True,
        margins_name="Grand Total"
    )
    
    # Reset index to make Department a column
    pivot_df = pivot_df.reset_index()
    
    # Remove any unnamed columns (blank columns that might appear)
    pivot_df = pivot_df.loc[:, ~pivot_df.columns.str.contains('^Unnamed', na=False)]
    
    # Get the department column name (first column after reset_index)
    dept_col_name = pivot_df.columns[0]
    
    # Get all other columns (status columns + Grand Total)
    other_cols = [col for col in pivot_df.columns if col != dept_col_name]
    
    # Reorder columns: Department first, then status columns, then Grand Total
    # Sort other columns to put Grand Total at the end
    status_cols = [col for col in other_cols if col != 'Grand Total']
    status_cols_sorted = sorted(status_cols)  # Sort status columns alphabetically
    
    # Final column order: Department, Status columns, Grand Total
    if 'Grand Total' in other_cols:
        final_cols = [dept_col_name] + status_cols_sorted + ['Grand Total']
    else:
        final_cols = [dept_col_name] + status_cols_sorted
    
    # Reorder columns
    pivot_df = pivot_df[final_cols]
    
    return pivot_df


def _write_pivot_to_sheet(ws, pivot_df, title):
    """
    Write pivot table data to worksheet with styling.
    
    Args:
        ws: openpyxl worksheet object
        pivot_df: pandas DataFrame (pivot table)
        title: Sheet title
    """
    # Write headers
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx, col_name in enumerate(pivot_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = str(col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, row_data in enumerate(pivot_df.values, start=2):
        # Check if this is the Grand Total row (last row)
        is_grand_total_row = (row_idx == len(pivot_df) + 1)
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            if is_grand_total_row:
                # Style Grand Total row with bold font and different background
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            else:
                # Alternate row coloring for regular rows
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Auto-adjust column widths
    for col_idx, col_name in enumerate(pivot_df.columns, start=1):
        max_length = max(
            len(str(col_name)),
            pivot_df[col_name].astype(str).map(len).max() if len(pivot_df) > 0 else 0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)


def _add_participation_chart(ws, pivot_df):
    """
    Add a bar chart to the participation summary sheet with data labels.
    Excludes the Grand Total row but includes all columns.
    Optimized for readability with large datasets.
    
    Args:
        ws: openpyxl worksheet object
        pivot_df: pandas DataFrame (pivot table with Grand Total row)
    """
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.text import RichText
    
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Participation Summary by Department"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Department"
    
    # Set chart size with extra height so axis titles and legend have space
    chart.height = 12   # increased height to keep titles/legend outside bars
    chart.width = 16    # slightly wider for readability
    
    # Determine data range - exclude Grand Total row (last row)
    num_rows = len(pivot_df)
    num_cols = len(pivot_df.columns)
    
    # Exclude the last row (Grand Total row) from chart
    # Data: columns 2 to end, rows 1 (header) to num_rows (exclude last row)
    # Categories: column 1, rows 2 to num_rows (exclude last row)
    data = Reference(ws, min_col=2, min_row=1, max_row=num_rows, max_col=num_cols)
    cats = Reference(ws, min_col=1, min_row=2, max_row=num_rows)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # Hide category labels on the axis so only counts show on the bars
    chart.x_axis.tickLblPos = "none"
    chart.x_axis.tickLblSkip = 1  # Kept for compatibility; has no effect when labels are hidden
    
    # Add data labels to show only value counts on each bar
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True         # Show numeric values on chart
    chart.dataLabels.showCatName = False    # Do not repeat category/label text
    chart.dataLabels.showLegendKey = False  # Do not include legend keys in labels
    chart.dataLabels.showSerName = False    # Do not show series name (e.g., "Good")
    chart.dataLabels.position = 'outEnd'   # Position labels outside bars
    
    # Increase spacing between bars
    chart.gapWidth = 150  # Gap between bar groups
    chart.overlap = -10   # Slight overlap for grouped bars
    
    # Show legend so each series (status) is identified, while labels show only counts
    chart.legend.position = "b"   # place legend below the chart
    chart.legend.overlay = False  # ensure legend does not overlap the plot area
    
    # Position chart below the data (accounting for Grand Total row)
    ws.add_chart(chart, f"A{num_rows + 3}")

