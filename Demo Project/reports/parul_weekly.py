"""
Parul Weekly Report Generator

Generates the Parul Weekly report with the Overall Data sheet transformed
according to the specified business rules.
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties


def read_csv_with_encoding(file_path):
    """
    Read CSV file with automatic encoding detection and robust parsing.
    """
    encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252', 'windows-1252']
    delimiters = [',', ';', '\t', '|']

    for encoding in encodings:
        for delimiter in delimiters:
            try:
                df = pd.read_csv(
                    file_path,
                    encoding=encoding,
                    delimiter=delimiter,
                    quotechar='"',
                    on_bad_lines='skip',
                    engine='python',
                    skipinitialspace=True
                )
                if len(df) > 0:
                    return df
            except (UnicodeDecodeError, UnicodeError):
                continue
            except Exception:
                continue

    for encoding in encodings:
        try:
            df = pd.read_csv(
                file_path,
                encoding=encoding,
                on_bad_lines='skip',
                engine='python',
                sep=None,
                skipinitialspace=True
            )
            if len(df) > 0:
                return df
        except Exception:
            continue

    return pd.read_csv(
        file_path,
        encoding='utf-8',
        errors='replace',
        on_bad_lines='skip',
        engine='python',
        sep=None,
        skipinitialspace=True
    )


def generate_parul_weekly_report(input_file, output_file):
    """
    Generate the Parul Weekly report.

    Args:
        input_file (str): Path to the input file (CSV, XLSX, or XLS)
        output_file (str): Path to save the processed report
    """
    df = _load_overall_data(input_file)
    processed_df = _process_overall_data(df)
    _write_formatted_workbook(processed_df, output_file)
    return output_file


def _load_overall_data(input_file):
    """
    Load the Overall Data sheet (or CSV) from the input file.
    """
    file_ext = input_file.lower().split('.')[-1]

    if file_ext == 'csv':
        return read_csv_with_encoding(input_file)

    try:
        return pd.read_excel(input_file, sheet_name='Overall Data')
    except ValueError:
        # If the sheet does not exist, load the first sheet
        return pd.read_excel(input_file)
    except Exception as e:
        # Attempt to read as CSV if Excel read fails
        try:
            return read_csv_with_encoding(input_file)
        except Exception:
            raise ValueError(f"Unable to read Overall Data sheet: {str(e)}")


def _process_overall_data(df):
    """
    Apply the Parul Weekly transformations to the Overall Data sheet.
    """
    df = df.copy()

    name_col = _get_column(df, ['name'])
    test_status_col = _get_column(df, ['test', 'status'])
    test_duration_col = _get_column(df, ['test', 'duration'])

    max_score_col = _get_column(df, ['max', 'score'])
    base_prefix = max_score_col[:max_score_col.lower().rfind('max score')].strip()
    student_score_col = _find_related_column(df, base_prefix, 'student score')
    total_percentage_col = _find_related_column(df, base_prefix, 'total percentage')

    # Portal Status column
    portal_status_col = 'Portal Status'
    df[portal_status_col] = df[name_col].apply(_compute_portal_status)

    # Attempt Status column
    attempt_status_col = 'Attempt Status'
    df[attempt_status_col] = df[test_duration_col].apply(_compute_attempt_status)

    # Order of key columns
    key_columns = [
        test_status_col,
        portal_status_col,
        attempt_status_col,
        max_score_col,
        student_score_col,
        total_percentage_col,
    ]

    # Category column (placed after total percentage)
    category_col = 'Category'
    df[category_col] = df[total_percentage_col].apply(_categorize_performance)
    key_columns.append(category_col)

    ordered_cols = [col for col in key_columns if col in df.columns]
    remaining_cols = [col for col in df.columns if col not in ordered_cols]
    df = df[ordered_cols + remaining_cols]

    return df


def _compute_portal_status(value):
    value_str = str(value).strip()
    return 'Not Activated' if value_str == '-' else 'Activated'


def _compute_attempt_status(value):
    value_str = '' if pd.isna(value) else str(value).strip()
    if value_str in ('', '-'):
        return 'No Attempt'
    if value_str == '0:00:00':
        return 'Unsuccessful Attempt'
    return 'Successful Attempt'


def _categorize_performance(value):
    if pd.isna(value):
        return 'Not Started'

    value_str = str(value).strip().replace('%', '')
    if value_str in ('', '-'):
        return 'Not Started'

    try:
        score = float(value_str)
        if score > 1:
            score = score / 100.0
    except ValueError:
        return 'Invalid Score'

    if score >= 0.75:
        return 'Good (75%+)'
    if score >= 0.50:
        return 'Satisfactory (50% - 75%)'
    if score >= 0.25:
        return 'Needs Attention (25% - 50%)'
    if score >= 0:
        return 'Intervention (0% - 25%)'
    return 'Invalid Score'


def _write_formatted_workbook(df, output_file):
    """
    Write the processed DataFrame to Excel and apply formatting.
    """
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Overall Data', index=False)

    wb = load_workbook(output_file)
    ws = wb['Overall Data']

    header_fill = PatternFill(start_color="1E4E79", end_color="1E4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style='thin', color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Header formatting
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    # Body formatting and borders
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Center-align key columns
    key_columns = {'Portal Status', 'Attempt Status', 'Category'}
    for col_name in df.columns:
        if col_name in key_columns or 'test status' in col_name.lower():
            col_idx = df.columns.get_loc(col_name) + 1
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = center_alignment

    # Auto-adjust column widths
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                continue
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    # Add Div Wise Performance Summary sheet
    _add_div_wise_performance_summary(wb, df)
    # Add Overall Performance Summary sheet
    _add_overall_performance_summary(wb, df)
    # Add Div Wise Participation Summary sheet
    _add_div_wise_participation_summary(wb, df)
    # Add Overall Participation Summary sheet
    _add_overall_participation_summary(wb, df)
    # Add Attempt Status Summary sheet
    _add_attempt_status_summary(wb, df)

    wb.save(output_file)


def _get_column(df, keywords):
    """
    Find the first column whose name contains all provided keywords.
    """
    for col in df.columns:
        col_lower = str(col).lower()
        if all(keyword in col_lower for keyword in keywords):
            return col
    raise ValueError(f"Required column containing keywords {keywords} not found.")


def _match_exact_column(df, target_col_name):
    """
    Find a column that matches the target name case-insensitively.
    """
    target_lower = target_col_name.lower()
    for col in df.columns:
        if str(col).lower() == target_lower:
            return col
    raise ValueError(f"Expected column '{target_col_name}' not found in the dataset.")


def _find_related_column(df, base_prefix, suffix):
    """
    Find a column that shares the base prefix and ends with the given suffix.
    """
    base_lower = (base_prefix or '').strip().lower()
    suffix_lower = suffix.strip().lower()
    exact_name = f"{base_prefix} {suffix}".strip().lower()

    # First, look for an exact match (with or without extra spaces)
    for col in df.columns:
        col_lower = str(col).strip().lower()
        if col_lower == exact_name:
            return col

    candidates = []
    for col in df.columns:
        col_lower = str(col).strip().lower()
        if col_lower.endswith(suffix_lower):
            leading = col_lower[:-len(suffix_lower)].rstrip()
            if base_lower and leading == base_lower:
                return col
            if base_lower and base_lower in leading:
                candidates.append(col)
            elif not base_lower:
                candidates.append(col)

    if candidates:
        return candidates[0]

    raise ValueError(f"Expected column matching '{base_prefix} {suffix}' not found in the dataset.")


def _add_div_wise_performance_summary(wb, df):
    """
    Create the Div Wise Performance Summary sheet with pivot and chart.
    """
    sheet_name = 'Div Wise Performance Summary'
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    dept_col = _get_column(df, ['department'])
    category_col = _match_exact_column(df, 'Category')
    count_col = _get_count_column(df)

    category_order = [
        'Good (75%+)',
        'Satisfactory (50% - 75%)',
        'Needs Attention (25% - 50%)',
        'Intervention (0% - 25%)',
        'Not Started',
    ]

    pivot_df = pd.pivot_table(
        df,
        index=dept_col,
        columns=category_col,
        values=count_col,
        aggfunc='count',
        fill_value=0
    )

    pivot_df = pivot_df.reindex(columns=category_order, fill_value=0)
    pivot_df.loc['Grand Total'] = pivot_df.sum()
    pivot_df.reset_index(inplace=True)

    _write_pivot_sheet(ws, pivot_df)
    _add_performance_chart(ws, pivot_df, category_order)


def _get_count_column(df):
    """
    Determine which column to use for counting rows in the pivot.
    """
    preferred_columns = ['Email', 'email', 'Name', 'name']
    for col in preferred_columns:
        if col in df.columns:
            return col
        for df_col in df.columns:
            if str(df_col).lower() == col.lower():
                return df_col
    return df.columns[0]


def _write_pivot_sheet(ws, pivot_df):
    """
    Write pivot data into the worksheet with formatting.
    """
    header_fill = PatternFill(start_color="1E4E79", end_color="1E4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style='thin', color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col_idx, col_name in enumerate(pivot_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    total_row_idx = len(pivot_df) + 1
    for row_idx, row_data in enumerate(pivot_df.values, start=2):
        is_total_row = (row_idx == total_row_idx)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_alignment
            cell.border = thin_border
            if is_total_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                continue
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)


def _add_performance_chart(ws, pivot_df, category_order):
    """
    Add a multi-series bar chart for department-wise performance.
    """
    num_rows = len(pivot_df)

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Department Wise Performance Summary"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Department"
    chart.height = 25
    chart.width = 40

    data = Reference(ws, min_col=2, min_row=1, max_row=num_rows + 1, max_col=len(category_order) + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=num_rows)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList(showVal=True)
    chart.gapWidth = 120
    chart.overlap = -10

    ws.add_chart(chart, f"A{num_rows + 4}")


def _add_overall_performance_summary(wb, df):
    """
    Create the Overall Performance Summary sheet with summary table and chart.
    """
    sheet_name = 'Overall Performance Summary'
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    category_col = _match_exact_column(df, 'Category')
    count_col = _get_count_column(df)

    category_order = [
        'Good (75%+)',
        'Satisfactory (50% - 75%)',
        'Needs Attention (25% - 50%)',
        'Intervention (0% - 25%)',
        'Not Started',
    ]

    summary_series = df.groupby(category_col)[count_col].count()
    summary_df = summary_series.reindex(category_order, fill_value=0).reset_index()
    summary_df.columns = ['Category', 'Count']
    grand_total = summary_df['Count'].sum()
    grand_row = pd.DataFrame([{'Category': 'Grand Total', 'Count': grand_total}])
    summary_df = pd.concat([summary_df, grand_row], ignore_index=True)

    table_rows = _write_overall_summary_table(ws, summary_df)
    _add_overall_summary_chart(ws, table_rows)


def _write_overall_summary_table(ws, summary_df):
    """
    Write the overall summary table with formatting.
    """
    header_fill = PatternFill(start_color="1E4E79", end_color="1E4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style='thin', color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    headers = list(summary_df.columns)
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    total_rows = len(summary_df) + 1  # includes header row
    for row_idx, row in enumerate(summary_df.itertuples(index=False, name=None), start=2):
        is_total_row = (row[0] == 'Grand Total')
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_alignment
            cell.border = thin_border
            if is_total_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Auto column widths
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                length = len(str(cell.value)) if cell.value is not None else 0
                max_length = max(max_length, length)
            except Exception:
                continue
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    return total_rows


def _add_overall_summary_chart(ws, total_rows):
    """
    Add a bar chart for overall performance distribution.
    """
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Overall Performance Distribution"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Category"
    chart.height = 20
    chart.width = 40

    # total_rows includes header + data rows (including Grand Total).
    data_end_row = total_rows - 1  # Exclude Grand Total
    data = Reference(ws, min_col=2, min_row=1, max_row=data_end_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=data_end_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList(showVal=True)

    ws.add_chart(chart, f"A{total_rows + 2}")


def _add_attempt_status_summary(wb, df):
    """
    Create the Attempt Status Summary sheet with table and chart.
    """
    sheet_name = 'Attempt Status Summary'
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    status_col = _get_column(df, ['attempt', 'status'])
    count_col = _get_count_column(df)

    status_order = [
        'Successful Attempt',
        'Unsuccessful Attempt',
        'No Attempt',
    ]

    summary_series = df.groupby(status_col)[count_col].count()
    summary_df = summary_series.reindex(status_order, fill_value=0).reset_index()
    summary_df.columns = ['Attempt Status', 'Count of Email']
    grand_total = summary_df['Count of Email'].sum()
    grand_row = pd.DataFrame([{'Attempt Status': 'Grand Total', 'Count of Email': grand_total}])
    summary_df = pd.concat([summary_df, grand_row], ignore_index=True)

    total_rows = _write_attempt_status_table(ws, summary_df)
    _add_attempt_status_chart(ws, total_rows)


def _write_attempt_status_table(ws, summary_df):
    """
    Write Attempt Status summary table with formatting.
    """
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(color="000000", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style='thin', color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    headers = list(summary_df.columns)
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    for row_idx, row in enumerate(summary_df.itertuples(index=False, name=None), start=2):
        is_total_row = (row[0] == 'Grand Total')
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_alignment
            cell.border = thin_border
            if is_total_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E9EDF5", end_color="E9EDF5", fill_type="solid")

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                length = len(str(cell.value)) if cell.value is not None else 0
                max_length = max(max_length, length)
            except Exception:
                continue
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    return len(summary_df) + 1


def _add_attempt_status_chart(ws, total_rows):
    """
    Add bar chart for Attempt Status distribution.
    """
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Attempt Status Distribution"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Attempt Status"
    chart.height = 20
    chart.width = 40

    # Exclude Grand Total row from chart
    data = Reference(ws, min_col=2, min_row=1, max_row=total_rows - 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=total_rows - 1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList(showVal=True)

    if chart.series:
        chart.series[0].graphicalProperties = GraphicalProperties(solidFill="1E4E79")

    ws.add_chart(chart, f"A{total_rows + 2}")


def _add_div_wise_participation_summary(wb, df):
    """
    Create the Div Wise Participation Summary sheet with table and chart.
    """
    sheet_name = 'Div Wise Participation Summary'
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    dept_col = _get_column(df, ['department'])
    status_col = _get_column(df, ['test', 'status'])
    count_col = _get_count_column(df)

    status_order = ['Completed', 'Not Started']

    pivot_df = pd.pivot_table(
        df,
        index=dept_col,
        columns=status_col,
        values=count_col,
        aggfunc='count',
        fill_value=0
    )
    pivot_df = pivot_df.reindex(columns=status_order, fill_value=0)
    pivot_df['Grand Total'] = pivot_df.sum(axis=1)
    pivot_df.loc['Grand Total'] = pivot_df.sum()
    pivot_df.reset_index(inplace=True)

    _write_participation_table(ws, pivot_df)
    _add_participation_chart(ws, len(pivot_df), len(status_order))


def _write_participation_table(ws, pivot_df):
    """
    Write participation summary table with formatting.
    """
    header_fill = PatternFill(start_color="1E4E79", end_color="1E4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style='thin', color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col_idx, col_name in enumerate(pivot_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    total_row_idx = len(pivot_df) + 1
    for row_idx, row in enumerate(pivot_df.itertuples(index=False, name=None), start=2):
        is_total_row = (row[0] == 'Grand Total')
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_alignment
            cell.border = thin_border
            if is_total_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                length = len(str(cell.value)) if cell.value is not None else 0
                max_length = max(max_length, length)
            except Exception:
                continue
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)


def _add_participation_chart(ws, total_rows, status_count):
    """
    Add a clustered bar chart for participation summary.
    """
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Department Wise Participation"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Department"
    chart.height = 25
    chart.width = 60

    # total_rows includes header and data rows (including Grand Total)
    data_end_row = total_rows  # includes Grand Total; keep for totals?
    # exclude Grand Total row from categories
    cat_end_row = total_rows - 1

    data = Reference(ws, min_col=2, min_row=1, max_row=data_end_row, max_col=status_count + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=cat_end_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList(showVal=True)
    chart.gapWidth = 100
    chart.overlap = -10

    ws.add_chart(chart, f"A{total_rows + 2}")


def _add_overall_participation_summary(wb, df):
    """
    Create the Overall Participation Summary sheet with table and chart.
    """
    sheet_name = 'Overall Participation Summary'
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    status_col = _get_column(df, ['test', 'status'])
    count_col = _get_count_column(df)

    status_order = ['Completed', 'Not Started']
    summary_series = df.groupby(status_col)[count_col].count()
    summary_df = summary_series.reindex(status_order, fill_value=0).reset_index()
    summary_df.columns = ['Test Status', 'Count']

    grand_total = summary_df['Count'].sum()
    grand_row = pd.DataFrame([{'Test Status': 'Grand Total', 'Count': grand_total}])
    summary_df = pd.concat([summary_df, grand_row], ignore_index=True)

    total_rows = _write_overall_participation_table(ws, summary_df)
    _add_overall_participation_chart(ws, total_rows)


def _write_overall_participation_table(ws, summary_df):
    """
    Write the overall participation table with formatting.
    """
    header_fill = PatternFill(start_color="1E4E79", end_color="1E4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style='thin', color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    headers = list(summary_df.columns)
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    for row_idx, row in enumerate(summary_df.itertuples(index=False, name=None), start=2):
        is_total_row = (row[0] == 'Grand Total')
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_alignment
            cell.border = thin_border
            if is_total_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                length = len(str(cell.value)) if cell.value is not None else 0
                max_length = max(max_length, length)
            except Exception:
                continue
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    return len(summary_df) + 1  # include header row


def _add_overall_participation_chart(ws, total_rows):
    """
    Add a bar chart for overall participation summary.
    """
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Summary"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Test Status"
    chart.height = 20
    chart.width = 40

    # Exclude the Grand Total row from chart data
    data = Reference(ws, min_col=2, min_row=1, max_row=total_rows - 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=total_rows - 1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList(showVal=True)

    # Apply blue color to bars
    if chart.series:
        chart.series[0].graphicalProperties = GraphicalProperties(solidFill="1E4E79")

    ws.add_chart(chart, f"A{total_rows + 2}")


