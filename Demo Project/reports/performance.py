"""
Performance Report Generator

This module generates a performance report from an Excel file.
It categorizes performance scores and creates pivot tables and charts.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reports.participation import generate_participation_report


def read_csv_with_encoding(file_path):
    """
    Read CSV file with automatic encoding detection and robust parsing.
    Tries multiple encodings and delimiters to handle various CSV formats.
    
    Args:
        file_path (str): Path to the CSV file
    
    Returns:
        pandas.DataFrame: DataFrame with CSV data
    
    Raises:
        ValueError: If file cannot be read with any method
    """
    encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252', 'windows-1252']
    delimiters = [',', ';', '\t', '|']
    
    # Try different combinations of encoding and delimiter
    for encoding in encodings:
        for delimiter in delimiters:
            try:
                df = pd.read_csv(
                    file_path,
                    encoding=encoding,
                    delimiter=delimiter,
                    quotechar='"',
                    on_bad_lines='skip',  # Skip malformed lines
                    engine='python',  # Use Python engine for better error handling
                    skipinitialspace=True
                )
                # If we got here and have data, return it
                if len(df) > 0:
                    return df
            except (UnicodeDecodeError, UnicodeError):
                continue
            except Exception as e:
                # If it's not an encoding error, try next delimiter
                if 'codec' in str(e).lower() or 'decode' in str(e).lower():
                    continue
                # For other errors, try next delimiter
                continue
    
    # If all combinations failed, try with most lenient settings (auto-detect separator)
    for encoding in encodings:
        try:
            df = pd.read_csv(
                file_path,
                encoding=encoding,
                on_bad_lines='skip',
                engine='python',
                sep=None,  # Auto-detect separator
                skipinitialspace=True
            )
            if len(df) > 0:
                return df
        except Exception:
            continue
    
    # Last resort: try with errors='replace'
    try:
        return pd.read_csv(
            file_path,
            encoding='utf-8',
            errors='replace',
            on_bad_lines='skip',
            engine='python',
            sep=None,
            skipinitialspace=True
        )
    except Exception as e:
        raise ValueError(f"Unable to read CSV file. Error: {str(e)}")


def generate_performance_report(input_file, output_file):
    """
    Generate a performance report with categorization and pivot tables.
    
    Args:
        input_file (str): Path to the input Excel file
        output_file (str): Path to save the output Excel file
    
    Returns:
        str: Path to the generated output file
    """
    # First generate participation report
    temp_output = output_file.replace('.xlsx', '_temp.xlsx')
    generate_participation_report(input_file, temp_output)
    
    # Load the workbook
    wb = load_workbook(temp_output)
    
    # Detect file extension and read original data accordingly
    file_ext = input_file.lower().split('.')[-1]
    
    if file_ext == "csv":
        df = read_csv_with_encoding(input_file)
    elif file_ext in ["xlsx", "xls"]:
        try:
            df = pd.read_excel(input_file)
        except Exception as e:
            error_msg = str(e).lower()
            # Check if it's a corrupted/invalid Excel file - try CSV as fallback
            if 'cannot be used in worksheets' in error_msg or 'badzipfile' in error_msg or 'corrupt' in error_msg:
                # Try reading as CSV instead (file might be misnamed)
                try:
                    df = read_csv_with_encoding(input_file)
                    # Successfully read as CSV, continue processing
                except:
                    raise ValueError(f"Invalid or corrupted file. The file '{input_file}' cannot be read as Excel or CSV. Please ensure the file is valid.")
            else:
                raise ValueError(f"Error reading Excel file: {str(e)}. Please ensure the file is valid.")
    else:
        raise ValueError("Unsupported file format. Please upload CSV or Excel files.")
    
    # Find Test Status column and percentage column
    status_col = None
    percentage_col = None
    status_idx = None
    percentage_idx = None
    
    # Find Test Status column
    for idx, col in enumerate(df.columns):
        col_lower = str(col).lower()
        if 'test status' in col_lower or ('status' in col_lower and 'test' in col_lower):
            status_col = col
            status_idx = idx
            break
    
    if status_col is None:
        raise ValueError("Required column 'Test Status' not found in the Excel file")
    
    # Find the FIRST column after Test Status that contains EXACTLY "total percentage"
    for idx in range(status_idx + 1, len(df.columns)):
        col = df.columns[idx]
        col_lower = str(col).lower()
        # Check for EXACT phrase "total percentage" (case insensitive)
        if 'total percentage' in col_lower:
            percentage_col = col
            percentage_idx = idx
            break
    
    if percentage_col is None:
        raise ValueError("Column containing 'total percentage' not found after 'Test Status' column")
    
    # Calculate Category column
    category_series = df[percentage_col].apply(_categorize_performance)
    
    # Insert Category column right after the percentage column
    # Get all columns up to and including percentage column
    cols_before = list(df.columns[:percentage_idx + 1])
    cols_after = list(df.columns[percentage_idx + 1:])
    
    # Create new column order with Category inserted
    new_columns = cols_before + ['Category'] + cols_after
    df = df.reindex(columns=new_columns)
    df['Category'] = category_series
    
    # Create performance pivot table
    dept_col = None
    for col in df.columns:
        if 'department' in str(col).lower():
            dept_col = col
            break
    
    if dept_col is None:
        raise ValueError("Required column 'Department' not found in the Excel file")
    
    # Find Name column for counting (or use first column)
    name_col = None
    for col in df.columns:
        if 'name' in str(col).lower():
            name_col = col
            break
    
    if name_col is None:
        # Use first column as fallback
        name_col = df.columns[0]
    
    # Create pivot table with Grand Total row and column
    pivot_df = pd.pivot_table(
        df,
        index=dept_col,
        columns='Category',
        values=name_col,
        aggfunc='count',
        fill_value=0,
        margins=True,
        margins_name="Grand Total"
    )
    
    # Reset index to make Department a column
    pivot_df = pivot_df.reset_index()
    
    # Remove any unnamed columns (blank columns that might appear)
    pivot_df = pivot_df.loc[:, ~pivot_df.columns.str.contains('^Unnamed', na=False)]
    
    # Get the department column name (first column after reset_index)
    dept_col_name = pivot_df.columns[0]
    
    # Get all other columns (category columns + Grand Total)
    other_cols = [col for col in pivot_df.columns if col != dept_col_name]
    
    # Reorder columns: Department first, then category columns, then Grand Total
    # Sort other columns to put Grand Total at the end
    category_cols = [col for col in other_cols if col != 'Grand Total']
    # Define preferred order for categories
    category_order = ['Good', 'Satisfactory', 'Need Attention', 'Intervention', 'Not Attended']
    category_cols_sorted = sorted(category_cols, key=lambda x: (
        category_order.index(x) if x in category_order else 999
    ))
    
    # Final column order: Department, Category columns, Grand Total
    if 'Grand Total' in other_cols:
        final_cols = [dept_col_name] + category_cols_sorted + ['Grand Total']
    else:
        final_cols = [dept_col_name] + category_cols_sorted
    
    # Reorder columns
    pivot_df = pivot_df[final_cols]
    
    # Create Performance Summary sheet
    ws_performance = wb.create_sheet("Performance Summary")
    _write_pivot_to_sheet(ws_performance, pivot_df, "Performance Summary")
    
    # Add bar chart
    _add_performance_chart(ws_performance, pivot_df)
    
    # Update Data sheet with Category column inserted in correct position
    ws_data = wb["Data"]
    _rewrite_data_sheet_with_category(ws_data, df)
    
    # Save final workbook
    wb.save(output_file)
    
    # Clean up temp file
    import os
    if os.path.exists(temp_output):
        os.remove(temp_output)
    
    return output_file


def _categorize_performance(value):
    """
    Categorize performance based on percentage value.
    
    Rules:
    > 75 → Good
    > 50 → Satisfactory
    > 25 → Need Attention
    <= 25 → Intervention
    NA / empty / "-" / "NA" / "n/a" / missing → Not Attended
    
    Args:
        value: Percentage value (numeric or string, may contain %)
    
    Returns:
        str: Category name
    """
    # Handle NaN, None, or empty values
    if pd.isna(value):
        return "Not Attended"
    
    # Convert to string and strip whitespace
    value_str = str(value).strip()
    
    # Handle empty strings and special values
    if value_str == '' or value_str.upper() in ['NA', 'N/A', '-', 'NULL', 'NONE']:
        return "Not Attended"
    
    try:
        # Remove % symbol if present
        value_str = value_str.replace('%', '').strip()
        
        # Convert to float
        num_value = float(value_str)
        
        # Apply category rules
        if num_value > 75:
            return "Good"
        elif num_value > 50:
            return "Satisfactory"
        elif num_value > 25:
            return "Need Attention"
        else:
            return "Intervention"
    except (ValueError, TypeError):
        # If conversion fails, treat as Not Attended
        return "Not Attended"


def _write_pivot_to_sheet(ws, pivot_df, title):
    """
    Write pivot table data to worksheet with styling.
    
    Args:
        ws: openpyxl worksheet object
        pivot_df: pandas DataFrame (pivot table)
        title: Sheet title
    """
    # Write headers
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx, col_name in enumerate(pivot_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = str(col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, row_data in enumerate(pivot_df.values, start=2):
        # Check if this is the Grand Total row (last row)
        is_grand_total_row = (row_idx == len(pivot_df) + 1)
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            if is_grand_total_row:
                # Style Grand Total row with bold font and different background
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            else:
                # Alternate row coloring for regular rows
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Auto-adjust column widths
    for col_idx, col_name in enumerate(pivot_df.columns, start=1):
        max_length = max(
            len(str(col_name)),
            pivot_df[col_name].astype(str).map(len).max() if len(pivot_df) > 0 else 0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)


def _add_performance_chart(ws, pivot_df):
    """
    Add a bar chart to the performance summary sheet with data labels.
    Excludes the Grand Total row but includes all columns.
    Optimized for readability with large datasets.
    
    Args:
        ws: openpyxl worksheet object
        pivot_df: pandas DataFrame (pivot table with Grand Total row)
    """
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.text import RichText
    
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Performance Summary by Department"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Department"
    
    # Set chart size with extra height so axis titles and legend have space
    chart.height = 12   # increased height to keep titles/legend outside bars
    chart.width = 16    # slightly wider for readability
    
    # Determine data range - exclude Grand Total row (last row)
    num_rows = len(pivot_df)
    num_cols = len(pivot_df.columns)
    
    # Exclude the last row (Grand Total row) from chart
    # Data: columns 2 to end, rows 1 (header) to num_rows (exclude last row)
    # Categories: column 1, rows 2 to num_rows (exclude last row)
    data = Reference(ws, min_col=2, min_row=1, max_row=num_rows, max_col=num_cols)
    cats = Reference(ws, min_col=1, min_row=2, max_row=num_rows)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # Hide category labels on the axis so only counts show on the bars
    chart.x_axis.tickLblPos = "none"
    chart.x_axis.tickLblSkip = 1  # Kept for compatibility; has no effect when labels are hidden
    
    # Add data labels to show only value counts on each bar
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True         # Show numeric values on chart
    chart.dataLabels.showCatName = False    # Do not repeat category/label text
    chart.dataLabels.showLegendKey = False  # Do not include legend keys in labels
    chart.dataLabels.showSerName = False    # Do not show series name (e.g., "Good")
    chart.dataLabels.position = 'outEnd'   # Position labels outside bars
    
    # Increase spacing between bars
    chart.gapWidth = 150  # Gap between bar groups
    chart.overlap = -10   # Slight overlap for grouped bars
    
    # Show legend so each series (category) is identified, while labels show only counts
    chart.legend.position = "b"   # place legend below the chart
    chart.legend.overlay = False  # ensure legend does not overlap the plot area
    
    # Position chart below the data (accounting for Grand Total row)
    ws.add_chart(chart, f"A{num_rows + 3}")


def _add_category_to_data_sheet(ws, df):
    """
    Add Category column to the Data sheet in the correct position.
    
    Args:
        ws: openpyxl worksheet object
        df: pandas DataFrame with Category column (already inserted in correct position)
    """
    # Find the Category column index in the DataFrame
    category_idx = list(df.columns).index('Category')
    category_col_num = category_idx + 1  # Excel columns are 1-indexed
    
    # Write Category header
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    category_cell = ws.cell(row=1, column=category_col_num)
    category_cell.value = "Category"
    category_cell.fill = header_fill
    category_cell.font = header_font
    category_cell.alignment = header_alignment
    category_cell.border = thin_border
    
    # Write Category data
    for row_idx, category in enumerate(df['Category'], start=2):
        cell = ws.cell(row=row_idx, column=category_col_num)
        cell.value = category
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Auto-adjust column width
    max_length = max(
        len("Category"),
        df['Category'].astype(str).map(len).max() if len(df) > 0 else 0
    )
    ws.column_dimensions[get_column_letter(category_col_num)].width = min(max_length + 2, 50)


def _rewrite_data_sheet_with_category(ws, df):
    """
    Rewrite the entire Data sheet with Category column in the correct position.
    
    Args:
        ws: openpyxl worksheet object
        df: pandas DataFrame with Category column already inserted in correct position
    """
    # Clear existing data (keep the sheet structure)
    ws.delete_rows(1, ws.max_row)
    
    # Header styling
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, row_data in enumerate(df.values, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            # Alternate row coloring
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Auto-adjust column widths
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_length = max(
            len(str(col_name)),
            df[col_name].astype(str).map(len).max() if len(df) > 0 else 0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

